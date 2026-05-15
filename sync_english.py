"""
Build ABERRATION_Dialogue_v5c.xlsx from dialogues_from_app.csv.
- CSV keys are NEVER modified (source of truth for Unity).
- English column is updated based on Russian (Russian = script authority).
- Excel shows original Unity key + friendly display name in separate column.
- Rows ordered by in-game chronology, not by speaker.
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE     = r"C:\Users\Disconnection\Desktop\GameDesign APP"
SRC_CSV  = os.path.join(BASE, "dialogues_from_app.csv")
OUT_CSV  = os.path.join(BASE, "dialogues_from_app.csv")   # only English updated
OUT_XLS  = os.path.join(BASE, "ABERRATION_Dialogue_v5c.xlsx")

# ── English translations keyed by ORIGINAL Unity/auto key ────────────────────
EN = {
    # 00. PROLOGUE
    "MEG_DAUGHTER_MESSAGE":
        "If you're watching this recording, please do me a favor. "
        "I know how the system works — that's why I'm reaching out to you directly. "
        "Get this message to my daughter.",
    "MEG_PR2_02":
        "Lily, Mom's in serious trouble and I'm not sure I'm coming back. "
        "Remember — no matter what happens, Mom is always with you. No matter what.",
    "STEVE_COMMENT_START":
        "Oh god... Another sob story about a mom and her little girl. How touching.",
    "JACK_PR2_01":
        "Meg, come in — are you okay? Did we lose the signal?",
    "MEG_PR2_03":
        "Yeah, all good. Had to leave a message. Don't worry about it.",
    "JACK_PR2_02":
        "Mm, got it. Alright — Baker Street should be right across from you, "
        "about seven minutes from there to the comm tower. I'm with you. Try to stay on.",
    "MEG_PR2_04":
        "Sorry. No problem.",
    # 01. ALLEY
    "MEG_FOUND_CHAIN":
        "Damn... Gotta find something to break this chain.",
    "MEG_PR4_01":
        "Locked. Need a key.",
    "MEG_PR5_01":
        "What the... There was a passage here!",
    "MEG_PR5_02":
        '"Souls that were meant to meet"... What the hell is that.',
    "MEG_PR5_03":
        "What are you doing in the trash, little guy...",
    "MEG_PR6_01":
        "Who's there...",
    "MEG_SCARE_REACTION":
        "What the fuck was that?!",
    "JACK_CHECK_OKAY":
        "Meg? Meg, are you in one piece? I can hear something going on.",
    "MEG_PR6_03":
        "I... yeah. Keeping moving.",
    "MEG_PR7_01":
        "Whose house?...",
    "MEG_PR7_02":
        "Damn. Damn damn damn.",
    "JACK_PR7_01":
        "Meg, what's going on?!",
    "MEG_BEING_FOLLOWED":
        "Someone's following me again!",
    "MEG_PR8_01":
        "Come on come on come on...",
    "MEG_PR8_02":
        "Damn!",
    "MEG_PR8_03":
        "...son of a bitch.",
    "JACK_PR8_01":
        "Meg? Meg, you made it out? I can hear you're outside. Breathe, breathe.",
    # 02. STREET
    "MEG_S02_B1_01":
        "Easy... deep breath in and out. Keep moving.",
    "MEG_S02_B1_02":
        "Jack, come in — I'm on the street.",
    "JACK_S02_B1_01":
        "Good, head up the street — there's a crossing to Baker Street, "
        "and from there it's just a short way to the tower.",
    "MEG_S02_B3_01":
        "Damn...",
    "MEG_S02_B4_01":
        "Come in, Jack — there's a blockage up ahead, street's blocked.",
    "JACK_S02_B4_01":
        "All good — there should be a crossing to the left, try getting through there.",
    "MEG_S02_B4_02":
        "Damn alleys...",
}

# ── Friendly display name per original key (Excel only, never touches CSV) ───
DISPLAY_NAME = {
    "MEG_DAUGHTER_MESSAGE":  "MEG_DAUGHTER_MESSAGE",         # already descriptive
    "MEG_PR2_02":            "MEG_ALTAR_MSG_TO_LILY",
    "STEVE_COMMENT_START":   "STEVE_COMMENT_START",          # already descriptive
    "JACK_PR2_01":           "JACK_CHECK_SIGNAL_LOST",
    "MEG_PR2_03":            "MEG_CONFIRM_ALL_CLEAR",
    "JACK_PR2_02":           "JACK_NAVIGATE_BAKER_ST",
    "MEG_PR2_04":            "MEG_ACKNOWLEDGE_MOVE",
    "MEG_FOUND_CHAIN":       "MEG_FOUND_CHAIN",
    "MEG_PR4_01":            "MEG_DOOR_LOCKED",
    "MEG_PR5_01":            "MEG_PASSAGE_GONE_SHOCK",
    "MEG_PR5_02":            "MEG_READ_SOULS_INSCRIPTION",
    "MEG_PR5_03":            "MEG_FIND_KEY_IN_TRASH",
    "MEG_PR6_01":            "MEG_HEAR_SOMEONE_WHISPER",
    "MEG_SCARE_REACTION":    "MEG_SCARE_REACTION",
    "JACK_CHECK_OKAY":       "JACK_CHECK_OKAY",
    "MEG_PR6_03":            "MEG_CONFIRM_OKAY_MOVING",
    "MEG_PR7_01":            "MEG_READ_WHOSE_HOUSE",
    "MEG_PR7_02":            "MEG_PANIC_CURSE",
    "JACK_PR7_01":           "JACK_DEMAND_SITUATION",
    "MEG_BEING_FOLLOWED":    "MEG_BEING_FOLLOWED",
    "MEG_PR8_01":            "MEG_PUSH_FORWARD_CHANT",
    "MEG_PR8_02":            "MEG_CURSE_HIT",
    "MEG_PR8_03":            "MEG_EXHAUSTED_CURSE",
    "JACK_PR8_01":           "JACK_CONFIRM_OUTSIDE_BREATHE",
    "MEG_S02_B1_01":         "MEG_CALM_DEEP_BREATH",
    "MEG_S02_B1_02":         "MEG_RADIO_ON_STREET",
    "JACK_S02_B1_01":        "JACK_NAVIGATE_UP_STREET",
    "MEG_S02_B3_01":         "MEG_SURREAL_DAMN",
    "MEG_S02_B4_01":         "MEG_REPORT_BLOCKAGE",
    "JACK_S02_B4_01":        "JACK_SUGGEST_LEFT_DETOUR",
    "MEG_S02_B4_02":         "MEG_DAMN_ALLEYS",
}

# ── Scene + Event labels per original key ────────────────────────────────────
ROW_META = {
    "MEG_DAUGHTER_MESSAGE":  ("00. PROLOGUE", "ALTAR & LILY'S MESSAGE"),
    "MEG_PR2_02":            ("00. PROLOGUE", "ALTAR & LILY'S MESSAGE"),
    "STEVE_COMMENT_START":   ("00. PROLOGUE", "ALTAR & LILY'S MESSAGE"),
    "JACK_PR2_01":           ("00. PROLOGUE", "ALTAR & LILY'S MESSAGE"),
    "MEG_PR2_03":            ("00. PROLOGUE", "ALTAR & LILY'S MESSAGE"),
    "JACK_PR2_02":           ("00. PROLOGUE", "ALTAR & LILY'S MESSAGE"),
    "MEG_PR2_04":            ("00. PROLOGUE", "ALTAR & LILY'S MESSAGE"),
    "MEG_FOUND_CHAIN":       ("01. ALLEY",    "ALLEY & CHAINED DOOR"),
    "MEG_PR4_01":            ("01. ALLEY",    "FIRST DESCENT TO BASEMENT"),
    "MEG_PR5_01":            ("01. ALLEY",    "SOULS THAT WERE MEANT TO MEET"),
    "MEG_PR5_02":            ("01. ALLEY",    "SOULS THAT WERE MEANT TO MEET"),
    "MEG_PR5_03":            ("01. ALLEY",    "SOULS THAT WERE MEANT TO MEET"),
    "MEG_PR6_01":            ("01. ALLEY",    "BOLT CUTTERS & PURSUER"),
    "MEG_SCARE_REACTION":    ("01. ALLEY",    "BOLT CUTTERS & PURSUER"),
    "JACK_CHECK_OKAY":       ("01. ALLEY",    "BOLT CUTTERS & PURSUER"),
    "MEG_PR6_03":            ("01. ALLEY",    "BOLT CUTTERS & PURSUER"),
    "MEG_PR7_01":            ("01. ALLEY",    "THIS IS MY HOUSE"),
    "MEG_PR7_02":            ("01. ALLEY",    "THIS IS MY HOUSE"),
    "JACK_PR7_01":           ("01. ALLEY",    "THIS IS MY HOUSE"),
    "MEG_BEING_FOLLOWED":    ("01. ALLEY",    "THIS IS MY HOUSE"),
    "MEG_PR8_01":            ("01. ALLEY",    "BREAKOUT TO THE STREET"),
    "MEG_PR8_02":            ("01. ALLEY",    "BREAKOUT TO THE STREET"),
    "MEG_PR8_03":            ("01. ALLEY",    "BREAKOUT TO THE STREET"),
    "JACK_PR8_01":           ("01. ALLEY",    "BREAKOUT TO THE STREET"),
    "MEG_S02_B1_01":         ("02. STREET",   "EXIT TO STREET & BAG MAZE"),
    "MEG_S02_B1_02":         ("02. STREET",   "EXIT TO STREET & BAG MAZE"),
    "JACK_S02_B1_01":        ("02. STREET",   "EXIT TO STREET & BAG MAZE"),
    "MEG_S02_B3_01":         ("02. STREET",   "SURREALISM & FALLING WARDROBE"),
    "MEG_S02_B4_01":         ("02. STREET",   "DEAD END & LOLLIPOP"),
    "JACK_S02_B4_01":        ("02. STREET",   "DEAD END & LOLLIPOP"),
    "MEG_S02_B4_02":         ("02. STREET",   "DEAD END & LOLLIPOP"),
}

# ── Chronological game order (original keys) ─────────────────────────────────
CHRON_ORDER = [
    "MEG_DAUGHTER_MESSAGE","MEG_PR2_02","STEVE_COMMENT_START",
    "JACK_PR2_01","MEG_PR2_03","JACK_PR2_02","MEG_PR2_04",
    "MEG_FOUND_CHAIN",
    "MEG_PR4_01",
    "MEG_PR5_01","MEG_PR5_02","MEG_PR5_03",
    "MEG_PR6_01","MEG_SCARE_REACTION","JACK_CHECK_OKAY","MEG_PR6_03",
    "MEG_PR7_01","MEG_PR7_02","JACK_PR7_01","MEG_BEING_FOLLOWED",
    "MEG_PR8_01","MEG_PR8_02","MEG_PR8_03","JACK_PR8_01",
    "MEG_S02_B1_01","MEG_S02_B1_02","JACK_S02_B1_01",
    "MEG_S02_B3_01",
    "MEG_S02_B4_01","JACK_S02_B4_01","MEG_S02_B4_02",
]
CHRON_IDX = {k: i for i, k in enumerate(CHRON_ORDER)}

SPEAKER_FILL      = {"Meg":"D6E4F0","Steve":"FCE5CD","Jack":"D9EAD3","Unknown":"E8DAEF"}
STATUS_FILL       = {"WRITTEN":"C6EFCE","DRAFT":"FFEB9C"}
STATUS_FONT_COLOR = {"WRITTEN":"276221","DRAFT":"9C6500"}
HDR_BG = "1A1A2E"
EVT_BG = "2A2A3E"

# KEY_UNITY = actual Unity key (never changes), KEY_NAME = friendly display name
COLUMNS = ["#","SCENE","EVENT","KEY_UNITY","KEY_NAME","TAG","SPEAKER","TONE OF VOICE",
           "ENGLISH","RUSSIAN","GERMAN","SPANISH","CHINESE","AUDIO FILE","STATUS","NOTE"]

def speaker_from_key(k):
    k = k.upper()
    if k.startswith("MEG_"):   return "Meg"
    if k.startswith("JACK_"):  return "Jack"
    if k.startswith("STEVE_"): return "Steve"
    return "Unknown"

# ── 1. Update English in CSV (keys untouched) ─────────────────────────────────
rows_in = []
with open(SRC_CSV, encoding="utf-8-sig", newline="") as f:
    reader = csv.DictReader(f)
    fieldnames = list(reader.fieldnames)
    rows_in = list(reader)

for r in rows_in:
    key = r.get("KEY","").strip()
    if key in EN:
        r["ENGLISH"] = EN[key]

with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows_in)
print(f"CSV updated (keys untouched): {OUT_CSV}")

# ── 2. Build Excel rows ───────────────────────────────────────────────────────
xl_rows = []
for r in rows_in:
    key = r.get("KEY","").strip()
    scene, event = ROW_META.get(key, ("??","??"))
    speaker = r.get("SPEAKER","").strip() or speaker_from_key(key)
    xl_rows.append({
        "KEY_UNITY":    key,
        "KEY_NAME":     DISPLAY_NAME.get(key, key),
        "SCENE":        scene,
        "EVENT":        event,
        "TAG":          r.get("TAG",""),
        "SPEAKER":      speaker,
        "TONE OF VOICE":r.get("TONE OF VOICE",""),
        "ENGLISH":      r.get("ENGLISH",""),
        "RUSSIAN":      r.get("RUSSIAN",""),
        "GERMAN":       r.get("GERMAN",""),
        "SPANISH":      r.get("SPANISH",""),
        "CHINESE":      r.get("CHINESE",""),
        "AUDIO FILE":   r.get("AUDIO FILE",""),
        "STATUS":       r.get("STATUS","DRAFT"),
        "NOTE":         r.get("NOTE",""),
    })

xl_rows.sort(key=lambda r: CHRON_IDX.get(r["KEY_UNITY"], 9999))
for i, r in enumerate(xl_rows, 1):
    r["#"] = i

# ── 3. Write Excel ────────────────────────────────────────────────────────────
def mkfill(c): return PatternFill("solid", fgColor=c)

wb = Workbook()
ws = wb.active
ws.title = "DIALOGUES"

hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
for ci, col in enumerate(COLUMNS, 1):
    c = ws.cell(row=1, column=ci, value=col)
    c.fill = mkfill(HDR_BG); c.font = hf
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

cur_scene = cur_event = None
ri = 2

for row in xl_rows:
    scene = row["SCENE"]; event = row["EVENT"]

    if scene != cur_scene:
        cur_scene = scene; cur_event = None
        c = ws.cell(row=ri, column=1, value=f"  {scene}")
        c.fill = mkfill(HDR_BG)
        c.font = Font(name="Arial", bold=True, color="CCCCFF", size=10)
        c.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(COLUMNS))
        ws.row_dimensions[ri].height = 22; ri += 1

    if event != cur_event:
        cur_event = event
        c = ws.cell(row=ri, column=1, value=f"     {event}")
        c.fill = mkfill(EVT_BG)
        c.font = Font(name="Arial", bold=True, color="AAAACC", size=9, italic=True)
        c.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(COLUMNS))
        ws.row_dimensions[ri].height = 18; ri += 1

    spk = row.get("SPEAKER","")
    sp  = mkfill(SPEAKER_FILL.get(spk, "FFFFFF"))
    st  = row.get("STATUS","DRAFT")
    stf = mkfill(STATUS_FILL.get(st, "FFFFFF"))
    stfc= STATUS_FONT_COLOR.get(st,"000000")

    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=ri, column=ci, value=row.get(col,""))
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col == "#":
            cell.font = Font(name="Consolas", size=8, color="888888")
            cell.fill = sp
            cell.alignment = Alignment(horizontal="center", vertical="top")
        elif col == "KEY_UNITY":
            # Actual Unity key — grey italic, smaller
            cell.font = Font(name="Consolas", size=8, italic=True, color="999999")
            cell.fill = sp
        elif col == "KEY_NAME":
            # Friendly display name — bold, prominent
            cell.font = Font(name="Consolas", size=9, bold=True)
            cell.fill = sp
        elif col == "STATUS":
            cell.fill = stf
            cell.font = Font(name="Arial", size=9, bold=True, color=stfc)
        elif col == "AUDIO FILE":
            cell.font = Font(name="Consolas", size=9); cell.fill = sp
        else:
            cell.fill = sp; cell.font = Font(name="Arial", size=9)
    ri += 1

widths = {"#":4,"SCENE":13,"EVENT":27,
          "KEY_UNITY":20,"KEY_NAME":30,
          "TAG":7,"SPEAKER":9,"TONE OF VOICE":26,
          "ENGLISH":42,"RUSSIAN":42,
          "GERMAN":28,"SPANISH":28,"CHINESE":18,
          "AUDIO FILE":16,"STATUS":10,"NOTE":28}
for ci, col in enumerate(COLUMNS, 1):
    ws.column_dimensions[get_column_letter(ci)].width = widths.get(col,14)

lg = wb.create_sheet("LEGEND")
lg["A1"] = "SPEAKER"; lg["A1"].font = Font(bold=True)
for i,(spk,color) in enumerate(SPEAKER_FILL.items(),2):
    lg.cell(row=i,column=1,value=spk)
    lg.cell(row=i,column=2,value="  ").fill = mkfill(color)
lg["D1"] = "HOW KEYS WORK"; lg["D1"].font = Font(bold=True)
lg.cell(row=2,column=4,value="KEY_UNITY = actual key Unity reads. NEVER rename.")
lg.cell(row=3,column=4,value="KEY_NAME  = human-readable label. Excel display only.")
lg.cell(row=4,column=4,value="WRITTEN   = approved, in Unity. DRAFT = not yet approved.")

wb.save(OUT_XLS)
print(f"Excel saved: {OUT_XLS}  ({len(xl_rows)} dialogue rows)")
